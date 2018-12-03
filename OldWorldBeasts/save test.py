from openpyxl import Workbook
from openpyxl import load_workbook

def loadData():
    try:
        load_workbook("data.xlsx")
    except:
        wbData = Workbook()
        wbData.save("data.xlsx")
