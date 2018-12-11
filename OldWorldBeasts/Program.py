# Made by Frederik Schrøder Koefoed

from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog
from easygui import enterbox
from tkinter import Toplevel
from tkinter import messagebox
import re
import random
import time

class AutocompleteEntry(Entry):
    def __init__(self, lista, *args, **kwargs):

        Entry.__init__(self, *args, **kwargs)
        self.lista = entryDic.keys()
        self.var = self["textvariable"]
        if self.var == '':
            self.var = self["textvariable"] = StringVar()

        self.var.trace('w', self.changed)
        self.bind("<Right>", self.selection)
        self.bind("<Up>", self.up)
        self.bind("<Down>", self.down)

        self.lb_up = False

    def changed(self, name, index, mode):

        if self.var.get() == '':
            self.lb.destroy()
            self.lb_up = False
        else:
            words = self.comparison()
            if words:
                if not self.lb_up:
                    self.lb = Listbox(width=46)
                    self.lb.configure()
                    self.lb.bind("<Double-Button-1>", self.selection)
                    self.lb.bind("<Right>", self.selection)
                    self.lb.place(x=10, y=73)
                    self.lb_up = True

                self.lb.delete(0, END)
                for w in words:
                    self.lb.insert(END,w)
            else:
                if self.lb_up:
                    self.lb.destroy()
                    self.lb_up = False

    def selection(self, event):

        if self.lb_up:
            self.var.set(self.lb.get(ACTIVE))
            self.lb.destroy()
            self.lb_up = False
            self.icursor(END)
            openByName(str(self.var.get()))

    def up(self, event):

        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != '0':
                self.lb.selection_clear(first=index)
                index = str(int(index)-1)
                self.lb.selection_set(first=index)
                self.lb.activate(index)

    def down(self, event):

        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != END:
                self.lb.selection_clear(first=index)
                index = str(int(index)+1)
                self.lb.selection_set(first=index)
                self.lb.activate(index)

    def comparison(self):
        pattern = re.compile('.*' + self.var.get() + '.*')
        return [w for w in self.lista if re.match(pattern, w)]

prompts = ["Navn (det der søges efter)", "Sidenummer (i WFRP Old World Beastiary)", "Weapon Skill (WS)", "Balistic Skill (BS)", "Strength (S)", "Toughness (T)",
"Agility (Ag)", "Intelligence (In)", "Will Power (WP)", "Fellowship (Fel)", "Attacks (A)", "Wounds (W)", "Movement (M)", "Magic (Mag)",
"Skills (separeret med et komma)", "Talents (separeret med et komma)", "Armour (None, Light, Medium, Heavy)", "Weapons (separeret med et komma)",
"Trappings (separeret med et komma)", "Special Rules (separeret med et komma)"]
entryCount = 0
nameData = ""
entryDic = {}

def selectData():
    global entryCount; global nameData; global entryDic
    entryCount = 0
    try:
        wbData = load_workbook("data.xlsx")
    except:
        wbData = Workbook()
        wbData.save("data.xlsx")
    wsData = wbData.active

    #updating entryCount
    x = 1
    while wsData["A" + str(x)].value != None:
        entryCount += 1
        entryDic[wsData["A" + str(x)].value] = entryCount
        x += 1
    textEntries = "Number of Entries: " + str(entryCount)
    vEntries.set(textEntries)

def newEntry(): # FIXME: check for same name = err
    global prompts; global entryCount; global directory; global nameData; global entryDic
    wbData = load_workbook("data.xlsx")
    wsData = wbData.active
    entryCount += 1
    promptCount = 0
    start = "A" + str(entryCount)
    slut = "T" + str(entryCount)
    for row in wsData[start:slut]:
        for cell in row:
            answer = enterbox(prompts[promptCount])
            print(answer)
            cell.value = answer.title()
            promptCount += 1
    entryDic[wsData[start].value] = entryCount
    wbData.save("data.xlsx")
    textEntries = "Number of Entries: " + str(entryCount)
    vEntries.set(textEntries)

def insertStatFrame(frame, row, column, color):
    name = Frame(frame, height=28, width=28)
    name.pack_propagate(0)
    name.grid(row=row, column=column, padx=(1, 1), pady=(1, 1))
    name.configure(bg=color)
    return name

def statLabel(frame, txt, bg):
    name = Label(frame, text=txt)
    name.place(x=14, y=14, anchor="center")
    name.configure(bg=bg)
    return name

def statText(frame, txt):
    name = Text(frame)
    name.tag_configure("center", justify='center')
    name.insert("1.0", txt)
    name.tag_add("center", "1.0", "end")
    name.pack()
    name.configure(font=("Helvetica", 9), bg="#A4A4A4")
    return name

def saveSpecial(txt, lineNR):
    wbData = load_workbook("Data.xlsx")
    wsData = wbData.active
    cell = "T" + str(lineNR)
    wsData[cell].value = txt
    wbData.save("data.xlsx")
    messagebox.showinfo("succes!", "The Special Rules have been saved!")

def openSpecial(name, txt, lineNR):
    windowS = Toplevel(root)
    windowS.title(name.title().strip() + ": Special Rules")
    xs = 304; ys = 286
    windowS.minsize(xs,ys)
    windowS.maxsize(xs,ys)
    windowS.configure(bg="#040030")

    extraFrame = Frame(windowS, height=7, width=1)
    extraFrame.pack()
    extraFrame.configure(bg="#040030")

    speT = Text(windowS, height=17, width=48)
    speT.insert(END, txt)
    speT.configure(font=("Helvetica", 8))
    speT.pack()

    saveb = Button(windowS, text="Save Special Rules", command=lambda: saveSpecial(speT.get(1.0, END), lineNR), height=1, width=40)
    saveb.pack(pady=3)

def saveAll(lineNR, namePage, ws, bs, s, t, ag, int, wp, fel, a, w, m, mag, skills, talents, trappings, armour, weapons):
    global entryDic
    wbData = load_workbook("data.xlsx")
    wsData = wbData.active
    #split name/pagenumber
    namePageList = namePage.split(":")
    name = list(entryDic.keys())[list(entryDic.values()).index(lineNR)]
    entryDic.pop(name, None)
    entryDic[namePageList[1].title().strip()] = lineNR
    wsData["A" + str(lineNR)].value = namePageList[1].title().strip()
    wsData["B" + str(lineNR)].value = namePageList[0].strip()
    wsData["C" + str(lineNR)].value = ws.strip()
    wsData["D" + str(lineNR)].value = bs.strip()
    wsData["E" + str(lineNR)].value = s.strip()
    wsData["F" + str(lineNR)].value = t.strip()
    wsData["G" + str(lineNR)].value = ag.strip()
    wsData["H" + str(lineNR)].value = int.strip()
    wsData["I" + str(lineNR)].value = wp.strip()
    wsData["J" + str(lineNR)].value = fel.strip()
    wsData["K" + str(lineNR)].value = a.strip()
    wsData["L" + str(lineNR)].value = w.strip()
    wsData["M" + str(lineNR)].value = m.strip()
    wsData["N" + str(lineNR)].value = mag.strip()
    wsData["O" + str(lineNR)].value = skills.strip().capitalize()
    wsData["P" + str(lineNR)].value = talents.strip().capitalize()
    wsData["Q" + str(lineNR)].value = trappings.strip().capitalize()
    armourList = armour.split(":")
    wsData["R" + str(lineNR)].value = armourList[1].strip().title()
    wsData["S" + str(lineNR)].value = weapons.strip().title()
    wbData.save("data.xlsx")
    messagebox.showinfo("succes!", "everything has been saved!")

def openByName(name):
    global entryDic
    random.seed(time.time())
    wbData = load_workbook("data.xlsx")
    wsData = wbData.active
    lineNR = int(entryDic[name])
    start = "A" + str(lineNR)
    slut = "T" + str(lineNR)
    cellList = []
    for row in wsData[start:slut]:
        for cell in row:
            cellList.append(str(cell.value))

    #UI
    window = Toplevel(root)
    window.title(name)
    xs = 450; ys = 200
    window.minsize(xs,ys)
    window.maxsize(xs,ys)

    leftFrame = Frame(window, height=ys, width=210)
    leftFrame.pack_propagate(0) # don't shrink
    leftFrame.pack(side=LEFT)

    nameFrame = Frame(leftFrame, height=ys-124, width=210)
    nameFrame.pack_propagate(0) # don't shrink
    nameFrame.pack(side=TOP)
    nameFrame.configure(bg="#2b2d2c")#040030

    nameT = Text(nameFrame, height=1, width=22)
    nameT.pack(pady=(9, 0))
    nameT.pack_propagate(0)
    nameTtxt = cellList[1].strip() + ": " + cellList[0].title().strip()
    nameT.insert(END, nameTtxt)
    nameT.configure(font=("Helvetica", 13))

    #initiative + special
    botNameFrame = Frame(nameFrame, height=(ys-124)/2, width=210)
    botNameFrame.pack_propagate(0) # don't shrink
    botNameFrame.pack(side=BOTTOM, padx=(4,0))
    botNameFrame.configure(bg="#2b2d2c")

    iniT = Text(botNameFrame, height=1, width=10)
    iniT.pack(padx=(2,1), side=LEFT)
    iniT.pack_propagate(0)
    iniT.insert(END, "Initiative: " + str(int(cellList[6]) + random.randint(1,10)))
    iniT.configure(font=("Helvetica", 13))

    specialb = Button(botNameFrame, text="Show Special", command=lambda: openSpecial(cellList[0], cellList[19], lineNR), height=1, width=13)
    specialb.pack(pady=1, padx=(1, 3), side=RIGHT)

    #Stats
    statFrame = Frame(leftFrame, height=124, width=210) #30x30 per box + 4 pixel in between
    statFrame.pack_propagate(0) # don't shrink
    statFrame.pack(side=BOTTOM)
    statFrame.configure(bg="#383a39") #1C1C1C

    wsTop = insertStatFrame(statFrame, 0, 0, "#6E6E6E")
    wsTopL = statLabel(wsTop, "WS", "#6E6E6E")
    bsTop = insertStatFrame(statFrame, 0, 1, "#6E6E6E")
    bsTopL = statLabel(bsTop, "BS", "#6E6E6E")
    sTop = insertStatFrame(statFrame, 0, 2, "#6E6E6E")
    sTopL = statLabel(sTop, "S", "#6E6E6E")
    tTop = insertStatFrame(statFrame, 0, 3, "#6E6E6E")
    tTopL = statLabel(tTop, "T", "#6E6E6E")
    agTop = insertStatFrame(statFrame, 0, 4, "#6E6E6E")
    agTopL = statLabel(agTop, "Ag", "#6E6E6E")
    intTop = insertStatFrame(statFrame, 0, 5, "#6E6E6E")
    intTopL = statLabel(intTop, "Int", "#6E6E6E")
    wpTop = insertStatFrame(statFrame, 0, 6, "#6E6E6E")
    wpTopL = statLabel(wpTop, "WP", "#6E6E6E")

    wsBot = insertStatFrame(statFrame, 1, 0, "#A4A4A4")
    wsBotT = statText(wsBot, cellList[2].strip())
    bsBot = insertStatFrame(statFrame, 1, 1, "#A4A4A4")
    bsBotT = statText(bsBot, cellList[3].strip())
    sBot = insertStatFrame(statFrame, 1, 2, "#A4A4A4")
    sBotT = statText(sBot, cellList[4].strip())
    tBot = insertStatFrame(statFrame, 1, 3, "#A4A4A4")
    tBotT = statText(tBot, cellList[5].strip())
    agBot = insertStatFrame(statFrame, 1, 4, "#A4A4A4")
    agBotT = statText(agBot, cellList[6].strip())
    intBot = insertStatFrame(statFrame, 1, 5, "#A4A4A4")
    intBotT = statText(intBot, cellList[7].strip())
    wpBot = insertStatFrame(statFrame, 1, 6, "#A4A4A4")
    wpBotT = statText(wpBot, cellList[8].strip())

    mellemStatFrame0 = Frame(statFrame, height=4, width=30)
    mellemStatFrame1 = Frame(statFrame, height=4, width=30)
    mellemStatFrame2 = Frame(statFrame, height=4, width=30)
    mellemStatFrame3 = Frame(statFrame, height=4, width=30)
    mellemStatFrame4 = Frame(statFrame, height=4, width=30)
    mellemStatFrame5 = Frame(statFrame, height=4, width=30)
    mellemStatFrame6 = Frame(statFrame, height=4, width=30)
    mellemStatFrame0.grid(row=2, column=0);
    mellemStatFrame1.grid(row=2, column=1);
    mellemStatFrame2.grid(row=2, column=2);
    mellemStatFrame3.grid(row=2, column=3);
    mellemStatFrame4.grid(row=2, column=4);
    mellemStatFrame5.grid(row=2, column=5);
    mellemStatFrame6.grid(row=2, column=6);
    mellemStatFrame0.configure(bg="#383a39")#383a39
    mellemStatFrame1.configure(bg="#383a39")
    mellemStatFrame2.configure(bg="#383a39")
    mellemStatFrame3.configure(bg="#383a39")
    mellemStatFrame4.configure(bg="#383a39")
    mellemStatFrame5.configure(bg="#383a39")
    mellemStatFrame6.configure(bg="#383a39")

    felTop = insertStatFrame(statFrame, 3, 6, "#6E6E6E")
    felTopL = statLabel(felTop, "Fel", "#6E6E6E")
    aTop = insertStatFrame(statFrame, 3, 0, "#6E6E6E")
    aTopL = statLabel(aTop, "A", "#6E6E6E")
    wTop = insertStatFrame(statFrame, 3, 1, "#6E6E6E")
    wTopL = statLabel(wTop, "W", "#6E6E6E")
    sbTop = insertStatFrame(statFrame, 3, 2, "#6E6E6E")
    sbTopL = statLabel(sbTop, "SB", "#6E6E6E")
    tbTop = insertStatFrame(statFrame, 3, 3, "#6E6E6E")
    tbTopL = statLabel(tbTop, "TB", "#6E6E6E")
    mTop = insertStatFrame(statFrame, 3, 4, "#6E6E6E")
    mTopL = statLabel(mTop, "M", "#6E6E6E")
    magTop = insertStatFrame(statFrame, 3, 5, "#6E6E6E")
    magTopL = statLabel(magTop, "Mag", "#6E6E6E")

    felBot = insertStatFrame(statFrame, 4, 6, "#A4A4A4")
    felBotT = statText(felBot, cellList[9].strip())
    aBot = insertStatFrame(statFrame, 4, 0, "#A4A4A4")
    aBotT = statText(aBot, cellList[10].strip())
    wBot = insertStatFrame(statFrame, 4, 1, "#A4A4A4")
    wBotT = statText(wBot, cellList[11].strip())
    sbBot = insertStatFrame(statFrame, 4, 2, "#A4A4A4")
    sbBotT = statText(sbBot, int(int(cellList[4].strip())/10))
    tbBot = insertStatFrame(statFrame, 4, 3, "#A4A4A4")
    tbBotT = statText(tbBot, int(int(cellList[5].strip())/10))
    mBot = insertStatFrame(statFrame, 4, 4, "#A4A4A4")
    mBotT = statText(mBot, cellList[12].strip())
    magBot = insertStatFrame(statFrame, 4, 5, "#A4A4A4")
    magBotT = statText(magBot, cellList[13].strip())

    #right side
    rightFrame = Frame(window, height=ys, width=240)
    rightFrame.pack(side=RIGHT)
    rightFrame.pack_propagate(0) # don't shrink
    rightFrame.configure(bg="#383a39")

    skillT = Text(rightFrame, height=6, width=58)
    skillT.pack(pady=(1, 0))
    skillT.pack_propagate(0)
    skillT.insert(END, cellList[14])
    skillT.configure(font=("Helvetica", 6))

    talentT = Text(rightFrame, height=6, width=58)
    talentT.pack(pady=(1, 0))
    talentT.pack_propagate(0)
    talentT.insert(END, cellList[15])
    talentT.configure(font=("Helvetica", 6))

    inventoryFrame = Frame(rightFrame, height=32, width=240)
    inventoryFrame.pack_propagate(0)
    inventoryFrame.configure(bg="#383a39")

    trappingsFrame = Frame(inventoryFrame, height=32, width=179)
    trappingsFrame.pack_propagate(0)
    trappingsFrame.configure(bg="#383a39")
    trappingsFrame.pack(side=LEFT)

    trappingsT = Text(trappingsFrame, height=3, width=44)
    trappingsT.pack(side=TOP)
    trappingsT.pack_propagate(0)
    trappingsT.insert(END, cellList[18])
    trappingsT.configure(font=("Helvetica", 6))

    armourFrame = Frame(inventoryFrame, height=32, width=61)
    armourFrame.pack_propagate(0)
    armourFrame.configure(bg="#383a39")
    armourFrame.pack(side=RIGHT)

    armourT = Text(armourFrame, height=3, width=10)
    armourT.pack(side=TOP)
    armourT.pack_propagate(0)
    armourT.insert(END, "Armour:\n" + cellList[16])
    armourT.configure(font=("Helvetica", 7))

    extraFrame = Frame(rightFrame, height=32, width=240)
    extraFrame.pack(pady=1, padx=1, side=BOTTOM)
    extraFrame.pack_propagate(0)
    extraFrame.configure(bg="#383a39")

    inventoryFrame.pack(padx=(2, 0), side=BOTTOM)

    weaponT = Text(extraFrame, height=2, width=35)
    weaponT.pack(pady=(1, 0), padx=(1, 0), side=LEFT)
    weaponT.pack_propagate(0)
    weaponT.insert(END, cellList[17])
    weaponT.configure(font=("Helvetica", 7))

    saveb = Button(extraFrame, text="Save", command=lambda: saveAll(lineNR, nameT.get(1.0,END), wsBotT.get(1.0,END), bsBotT.get(1.0,END), sBotT.get(1.0,END), tBotT.get(1.0,END),
    agBotT.get(1.0,END), intBotT.get(1.0,END), wpBotT.get(1.0,END), felBotT.get(1.0,END), aBotT.get(1.0,END), wBotT.get(1.0,END), mBotT.get(1.0,END), magBotT.get(1.0,END),
    skillT.get(1.0,END), talentT.get(1.0,END), trappingsT.get(1.0,END), armourT.get(1.0,END), weaponT.get(1.0,END)), height=1, width=6)
    saveb.pack(pady=1, padx=2, side=RIGHT)

#UI

root = Tk()
root.title("New Worlds Beastiary")
root.minsize(300,245)
root.maxsize(300,245)
rootColor = "#040030"

def insertMellemFrame(height, place):
    mellemFrame = Frame(place, height=height, width=300)
    mellemFrame.pack_propagate(0) # don't shrink
    mellemFrame.pack()
    mellemFrame.configure(bg=rootColor)

insertMellemFrame(5, root)

entryFrame = Frame(root, height=45, width=300)
entryFrame.pack_propagate(0) # don't shrink
entryFrame.pack()
entryFrame.configure(bg=rootColor)

vEntries = StringVar()
textEntries = "Number of Entries: " + str(entryCount)
vEntries.set(textEntries)

insertMellemFrame(5, root)

entryFillFrame = Frame(entryFrame, height=2, width=6)
entryFillFrame.configure(bg=rootColor)

entryl2 = Label(entryFrame, textvariable=vEntries)
entryl2.configure(bg=rootColor, fg="#dee0de", font=("Helvetica", 14))
entryb = Button(entryFrame, text="New Entry", command=newEntry, height=1, width=9)
entryb.configure(font=("Helvetica", 13), bg="#d1d1d1")

entryFillFrame.pack(side=LEFT)
entryb.pack(side=LEFT)
entryl2.pack(side=LEFT, padx=(6, 0))


autoFrame = Frame(root, height=200, width=300)
autoFrame.pack_propagate(0) # don't shrink
autoFrame.pack()
autoFrame.configure(bg=rootColor)

entry = AutocompleteEntry(entryDic, autoFrame)
entry.configure(width=46)
entry.pack()

selectData()

root.mainloop()
